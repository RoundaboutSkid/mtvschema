#!/usr/bin/env python3
"""
Build the Medeltidsveckan timetable UI from the fetched data file.

Reads ``medeltidsveckan_events.json`` (or a ``.csv``) produced by
``fetch_programme.py`` and renders:

- ``medeltidsveckan_schema.html`` – an interactive day-by-day timeline.
- ``medeltidsveckan_schema.xlsx`` – the same layout as an Excel workbook.

The timeline is a CSS grid per day. Each event is placed with
``grid-row: start / end`` so a multi-slot event is **one continuous block**
instead of being repeated per row, and events that overlap at the *same venue*
are split into side-by-side **lanes**.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from collections import defaultdict
from pathlib import Path

from medeltidsveckan_common import (
    DEFAULT_CATEGORY_COLOR,
    DEFAULT_EXCLUDE_FILENAME,
    DayLayout,
    Event,
    Placement,
    _layout_intervals,
    build_day_layout,
    category_color,
    dedupe,
    filter_events,
    load_events,
    load_exclude_patterns,
    minutes_to_hhmm,
)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    openpyxl = None


# --------------------------------------------------------------------------- #
# Schema data  (window.MV_DATA, consumed and rendered by frontend/app.js)
# --------------------------------------------------------------------------- #
PX_PER_MIN = 1.3          # vertical scale: 1 minute -> px
NO_CATEGORY = "__none__"  # sentinel for events without a category
CONFIG_FILENAME = "medeltidsveckan_config.json"  # persists e.g. the subscription endpoint
VENUES_FILENAME = "medeltidsveckan_venues.json"  # editable place→type/zone/map-point table
DEFAULT_VENUE_TYPE = "ovrigt"
DEFAULT_VENUE_ICON = "\U0001F4CD"  # 📍 fallback for places missing from the table
# Header colours for the zone-band view (one per zone id, parchment-friendly).
ZONE_COLORS = {"Z1": "B5793A", "Z2": "5B8C5A", "Z3": "5C7CA8", "Z4": "8E6BA8", "Z5": "8A6D3B"}
DEFAULT_ZONE_COLOR = "6B7280"
FRONTEND_DIR = Path(__file__).resolve().parent / "frontend"  # editable HTML/CSS/JS sources


def event_dom_id(e: Event) -> str:
    """Stable client-side id for an event.

    Must match the ``data-id`` attribute used in the HTML so that the browser,
    the print/ICS code and the subscription Worker all agree on the same key.
    """
    return hashlib.md5(
        f"{e.date}|{e.start}|{e.end}|{e.venue}|{e.title}|{e.organizer}".encode("utf-8")
    ).hexdigest()[:12]


def _pos(minute_offset: int) -> float:
    return round(minute_offset * PX_PER_MIN, 1)


def _event_record(p: Placement, day: DayLayout, lane_count: int) -> dict:
    """Build the JSON record app.js needs to recreate one ``.event`` block.

    All layout maths (position, lane width, tooltip, search text) happens here so
    the browser only maps values onto the DOM. Keys mirror the old ``data-*``
    attributes.
    """
    e = p.event
    color = category_color(e.category)
    top = _pos(e.start_min - day.day_start_min)
    height = _pos(e.end_min - e.start_min)
    left = round(p.lane / lane_count * 100, 4)
    width = round(p.span / lane_count * 100, 4)
    time_txt = f"{e.start}\u2013{e.end}"
    search = " ".join([e.title, e.organizer, e.venue, e.category, e.status, e.description]).lower()

    tooltip = f"{time_txt} \u00b7 {e.title}"
    if e.organizer:
        tooltip += f" \u00b7 {e.organizer}"
    tooltip += f" \u00b7 {e.venue}"
    if e.description:
        snippet = " ".join(e.description.split())
        if len(snippet) > 220:
            snippet = snippet[:220].rstrip() + "\u2026"
        tooltip += f"\n\n{snippet}"
    tooltip += "\n\n(Klicka för detaljer)"

    rec: dict[str, object] = {
        "id": event_dom_id(e), "s": e.start_min, "e": e.end_min,
        "top": top, "height": height, "left": left, "width": width,
        "color": color, "short": height < 34,
        "title": e.title, "time": time_txt, "venue": e.venue,
        "cat": e.category or "", "catKey": e.category if e.category else NO_CATEGORY,
        "search": search, "tooltip": tooltip,
    }
    if e.organizer:
        rec["org"] = e.organizer
        rec["orgShow"] = e.organizer.strip().lower() != e.venue.strip().lower()
    if e.status:
        rec["status"] = e.status
    if e.ticket_url:
        rec["ticket"] = e.ticket_url
    if e.description:
        rec["desc"] = e.description
    return rec


def _zone_placements(day_events: list[Event], zone_of: dict[str, str | None]):
    """Lane-pack a day's events grouped by *zone* (parallel to the venue layout).

    Returns ``{zone_id: (lane_count, [(event, lane, span)])}`` using the same
    interval-graph colouring as the venue layout, so the two views share the
    exact same time axis and only differ in horizontal grouping.
    """
    by_zone: dict[str, list[Event]] = defaultdict(list)
    for e in day_events:
        by_zone[zone_of.get(e.venue) or "Z?"].append(e)
    out: dict[str, tuple[int, list]] = {}
    for z, evs in by_zone.items():
        lane_count, placed = _layout_intervals([(e.start_min, e.end_min, e) for e in evs])
        out[z] = (max(lane_count, 1), placed)
    return out


def _day_record(layout: DayLayout, day_events: list[Event], venue_meta: dict[str, dict],
                zone_of: dict[str, str | None], zones_meta: list[dict]) -> dict:
    """Serialise one day's computed layout for ``window.MV_DATA``.

    Emits two parallel groupings sharing one time axis: ``venues`` (the flow
    view) and ``zones`` (the zone-band view). Each event record is created once
    and carries both its venue lane (``left``/``width``) and its zone lane
    (``zLeft``/``zWidth``) plus the place ``icon``.
    """
    track_h = _pos(layout.day_end_min - layout.day_start_min)
    slot_px = round(layout.slot_minutes * PX_PER_MIN, 1)
    ticks = []
    minute = layout.day_start_min
    while minute <= layout.day_end_min:
        ticks.append({
            "top": _pos(minute - layout.day_start_min),
            "label": minutes_to_hhmm(minute),
            "hour": minute % 60 == 0,
        })
        minute += layout.slot_minutes

    rec_of: dict[int, dict] = {}
    venues = []
    for v in layout.venues:
        evs = []
        for p in v.placements:
            rec = _event_record(p, layout, v.lane_count)
            rec["icon"] = venue_meta.get(rec["venue"], {}).get("icon", DEFAULT_VENUE_ICON)
            rec_of[id(p.event)] = rec
            evs.append(rec)
        venues.append({
            "venue": v.venue,
            "laneCount": v.lane_count,
            "minW": max(150, v.lane_count * 132),
            "events": evs,
        })

    # Zone grouping shares the same time axis; attach zone lane info to each record.
    zlay = _zone_placements(day_events, zone_of)
    zone_label = {z["id"]: z.get("label", z["id"]) for z in zones_meta}
    order = [z["id"] for z in zones_meta] + ["Z?"]
    zones = []
    for z in order:
        if z not in zlay:
            continue
        lane_count, placed = zlay[z]
        icons: list[str] = []
        for event, lane, span in placed:
            rec = rec_of.get(id(event))
            if rec is None:
                continue
            rec["zone"] = z
            rec["zLeft"] = round(lane / lane_count * 100, 4)
            rec["zWidth"] = round(span / lane_count * 100, 4)
            ic = rec.get("icon")
            if ic and ic not in icons:
                icons.append(ic)
        zones.append({
            "id": z,
            "label": zone_label.get(z, "Övrigt" if z == "Z?" else z),
            "color": ZONE_COLORS.get(z, DEFAULT_ZONE_COLOR),
            "icons": icons[:5],
            "laneCount": lane_count,
            "minW": max(180, lane_count * 132),
        })

    label = f"{(layout.weekday[:3] or layout.date)} {layout.date[5:]}"
    title = f"{layout.weekday} {layout.date}".strip()
    return {
        "date": layout.date, "weekday": layout.weekday, "label": label, "title": title,
        "dayStart": layout.day_start_min, "dayEnd": layout.day_end_min,
        "slot": layout.slot_minutes, "px": PX_PER_MIN, "slotPx": slot_px, "trackH": track_h,
        "ticks": ticks, "venues": venues, "zones": zones,
    }


def _load_venue_meta() -> dict:
    """Load the editable place→type/zone/map-point table next to the scripts.

    Returns ``{"types": {typ: icon}, "zones": [{"id","label"}…], "places":
    {venue: {"typ","icon","zon","punkt"}}}``. A missing or invalid file yields
    empty maps so the page still builds and the client falls back to a pin icon.
    """
    path = Path(__file__).resolve().parent / VENUES_FILENAME
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return {"types": {}, "zones": [], "places": {}}
    if not isinstance(data, dict):
        return {"types": {}, "zones": [], "places": {}}
    types = data.get("typer", {}) or {}
    zoner = data.get("zoner", {}) or {}
    platser = data.get("platser", {}) or {}
    zones = [{"id": k, "label": v} for k, v in zoner.items()]
    places: dict[str, dict] = {}
    for name, info in platser.items():
        if not isinstance(info, dict):
            continue
        typ = info.get("typ") or DEFAULT_VENUE_TYPE
        places[name] = {
            "typ": typ,
            "icon": types.get(typ, DEFAULT_VENUE_ICON),
            "zon": info.get("zon"),
            "punkt": info.get("punkt"),
        }
    return {"types": types, "zones": zones, "places": places}


def _data_payload(events: list[Event], slot_minutes: int, ics_endpoint: str) -> dict:
    """Whole-programme data the client renders from: filters + per-day layout."""
    by_day: dict[str, list[Event]] = defaultdict(list)
    for e in events:
        by_day[e.date].append(e)
    layouts = [build_day_layout(d, by_day[d], slot_minutes) for d in sorted(by_day)]

    cats = sorted({e.category for e in events if e.category})
    cat_items = [{"val": c, "label": c, "color": category_color(c)} for c in cats]
    if any(not e.category for e in events):
        cat_items.append({"val": NO_CATEGORY, "label": "Utan kategori", "color": DEFAULT_CATEGORY_COLOR})
    venues = sorted({e.venue for e in events}, key=lambda s: (s == "Okänd plats", s.lower()))

    meta = _load_venue_meta()
    places = meta["places"]
    venue_meta: dict[str, dict] = {}
    unknown: list[str] = []
    for v in venues:
        info = places.get(v)
        if info is None:
            unknown.append(v)
            venue_meta[v] = {"typ": DEFAULT_VENUE_TYPE, "icon": DEFAULT_VENUE_ICON,
                             "zon": None, "punkt": None}
        else:
            venue_meta[v] = info
    if unknown:
        print(f"OBS: {len(unknown)} plats(er) saknas i {VENUES_FILENAME} "
              f"(får standardikon {DEFAULT_VENUE_ICON}, ingen zon):")
        for v in unknown:
            print(f"  • {v}")

    zone_of = {v: (info.get("zon") if info else None) for v, info in venue_meta.items()}
    days = [_day_record(l, by_day[l.date], venue_meta, zone_of, meta["zones"]) for l in layouts]

    # Weekly-max lanes per venue -> a stable column width for the "all places"
    # view so a venue keeps the same x-position (and width) on every day.
    weekly_max: dict[str, int] = {}
    for l in layouts:
        for v in l.venues:
            weekly_max[v.venue] = max(weekly_max.get(v.venue, 1), v.lane_count)
    places = [{
        "venue": v,
        "icon": venue_meta.get(v, {}).get("icon", DEFAULT_VENUE_ICON),
        "laneCount": weekly_max.get(v, 1),
        "minW": max(132, weekly_max.get(v, 1) * 120),
        "minWc": max(72, weekly_max.get(v, 1) * 60),
    } for v in venues]

    return {
        "icsEndpoint": ics_endpoint or "",
        "cats": cat_items,
        "venues": venues,
        "venueMeta": venue_meta,
        "types": meta["types"],
        "zones": meta["zones"],
        "places": places,
        "days": days,
    }


def write_html(events: list[Event], path: Path, slot_minutes: int, ics_endpoint: str = "") -> None:
    """Render the single-file schema by inlining the ``frontend/`` templates.

    The page body, styles and behaviour now live in editable files
    (index.html, styles.css, app.js). Here we compute the layout into
    ``window.MV_DATA`` and splice the three sources into the shell, keeping the
    output a single, double-click-openable file.
    """
    payload = _data_payload(events, slot_minutes, ics_endpoint)
    data_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    # Keep the JSON safe to embed inside a <script> block.
    data_json = (data_json.replace("<", "\\u003c")
                 .replace(">", "\\u003e").replace("&", "\\u0026"))

    shell = (FRONTEND_DIR / "index.html").read_text(encoding="utf-8")
    css = (FRONTEND_DIR / "styles.css").read_text(encoding="utf-8")
    app = (FRONTEND_DIR / "app.js").read_text(encoding="utf-8")

    out = (
        shell
        .replace("/*__MV_STYLES__*/", css)
        .replace("/*__MV_DATA__*/", "window.MV_DATA = " + data_json + ";")
        .replace("/*__MV_APP__*/", app)
    )
    path.write_text(out, encoding="utf-8")


def write_worker_data(events: list[Event], path: Path) -> int:
    """Dump the programme as a JS module the subscription Worker can import.

    The Worker only stores each user's *favourite ids*; the event details live
    here, keyed by the same id as the HTML's ``data-id``. Keys are short to keep
    the bundle small: d(ate) s(tart_min) e(nd_min) t(itle) v(enue) o(rganizer)
    c(ategory) desc tk(ticket url).
    """
    programme: dict[str, dict] = {}
    for e in events:
        rec: dict[str, object] = {
            "d": e.date, "s": e.start_min, "e": e.end_min, "t": e.title, "v": e.venue,
        }
        if e.organizer:
            rec["o"] = e.organizer
        if e.category:
            rec["c"] = e.category
        if e.description:
            rec["desc"] = e.description
        if e.ticket_url:
            rec["tk"] = e.ticket_url
        programme[event_dom_id(e)] = rec
    payload = json.dumps(programme, ensure_ascii=False, separators=(",", ":"))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "// Auto-generated by build_schedule.py \u2013 do not edit by hand.\n"
        "// Maps event id (same as the schema's data-id) to event details.\n"
        f"export const PROGRAMME = {payload};\n",
        encoding="utf-8",
    )
    return len(programme)


# --------------------------------------------------------------------------- #
# Excel (mirrors the lane layout with merged cells)
# --------------------------------------------------------------------------- #
def _style_excel_event(cell, color: str) -> None:
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(color="FFFFFF", bold=False, size=9)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_excel(events: list[Event], path: Path, slot_minutes: int) -> None:
    if openpyxl is None:
        raise RuntimeError("Installera openpyxl: pip install openpyxl")

    by_day: dict[str, list[Event]] = defaultdict(list)
    for e in events:
        by_day[e.date].append(e)

    wb = openpyxl.Workbook()
    raw = wb.active
    raw.title = "Alla event"
    headers = ["date", "weekday", "start", "end", "venue", "title", "organizer",
               "category", "status", "ticket_url", "duration_source"]
    raw.append(headers)
    for e in sorted(events, key=lambda x: (x.date, x.start, x.venue, x.title)):
        raw.append([e.date, e.weekday, e.start, e.end, e.venue, e.title,
                    e.organizer, e.category, e.status, e.ticket_url, e.duration_source])
    _style_raw_sheet(raw)

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="5C2C14")
    header_font = Font(color="FFFFFF", bold=True)

    for date_s in sorted(by_day):
        layout = build_day_layout(date_s, by_day[date_s], slot_minutes)
        sheet_title = f"{(layout.weekday[:3] or '')} {date_s[5:]}".strip()
        ws = wb.create_sheet(sheet_title[:31])

        venue_base: dict[str, int] = {}
        col = 2
        ws.cell(1, 1, "Tid").fill = header_fill
        ws.cell(1, 1).font = header_font
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        for v in layout.venues:
            venue_base[v.venue] = col
            c1, c2 = col, col + v.lane_count - 1
            if c2 > c1:
                ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
            cell = ws.cell(1, c1, v.venue)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col += v.lane_count
        total_cols = col - 1

        # Time labels every slot; one spreadsheet row per base unit.
        rows_per_slot = layout.rows_per_slot
        for row_idx, minute in layout.slot_ticks():
            r = row_idx + 2
            tcell = ws.cell(r, 1, minutes_to_hhmm(minute))
            tcell.font = Font(bold=True, size=9, color="6B5D4F")
            tcell.alignment = Alignment(horizontal="right", vertical="top")

        # Event blocks: vertical merge for duration, horizontal merge for lane span.
        for v in layout.venues:
            for p in v.placements:
                r1 = p.row_start + 2
                r2 = p.row_end + 2 - 1
                c1 = venue_base[v.venue] + p.lane
                c2 = c1 + p.span - 1
                if r2 > r1 or c2 > c1:
                    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                text = f"{p.event.start}\u2013{p.event.end}\n{p.event.title}"
                if p.event.organizer and p.event.organizer.lower() != p.event.venue.lower():
                    text += f"\n{p.event.organizer}"
                if p.event.status:
                    text += f"\n[{p.event.status}]"
                cell = ws.cell(r1, c1, text)
                _style_excel_event(cell, category_color(p.event.category))

        # Column widths + row heights + borders on the time column.
        ws.column_dimensions[get_column_letter(1)].width = 7
        for c in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 22
        per_base_height = max(8.0, layout.base_minutes * 0.9)
        for r in range(2, layout.rows + 2):
            ws.row_dimensions[r].height = per_base_height
        ws.freeze_panes = "B2"
        for r in range(1, layout.rows + 2):
            ws.cell(r, 1).border = border

    wb.save(path)


def _style_raw_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="5C2C14")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {"date": 12, "weekday": 10, "start": 7, "end": 7, "venue": 30,
              "title": 42, "organizer": 26, "category": 16, "status": 14,
              "ticket_url": 40, "duration_source": 16}
    for idx, name in enumerate([c.value for c in ws[1]], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(name, 14)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #
def _load_config(outdir: Path) -> dict:
    """Load the small persisted config (e.g. the subscription endpoint)."""
    path = outdir / CONFIG_FILENAME
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except (ValueError, OSError):
            pass
    return {}


def _save_config(outdir: Path, cfg: dict) -> None:
    path = outdir / CONFIG_FILENAME
    path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _resolve_inputs(outdir: Path, explicit: list[str] | None,
                    include_inofficial: bool = True) -> list[Path]:
    """Pick the data file(s) to render.

    With ``--input`` the given files are used as-is. Otherwise the official
    programme file is auto-discovered in ``outdir`` and, unless disabled, the
    unofficial programme file (``medeltidsveckan_inofficial.*``) is merged in too.
    """
    if explicit:
        return [Path(x) for x in explicit]

    def first_existing(*names: str) -> Path | None:
        for name in names:
            candidate = outdir / name
            if candidate.exists():
                return candidate
        return None

    paths: list[Path] = []
    main = first_existing("medeltidsveckan_events.json", "medeltidsveckan_events.csv")
    if main:
        paths.append(main)
    if include_inofficial:
        inof = first_existing("medeltidsveckan_inofficial.json", "medeltidsveckan_inofficial.csv")
        if inof:
            paths.append(inof)
    if not paths:
        raise SystemExit(
            f"Hittade ingen datafil i {outdir}. Kör först: python fetch_programme.py"
        )
    return paths


def _collect_exclude_patterns(args, outdir: Path) -> list[str]:
    """Gather exclusion patterns from the exclude file (if any) and --exclude."""
    patterns: list[str] = []
    if not args.no_exclude_file:
        if args.exclude_file:
            exclude_path = Path(args.exclude_file)
        else:
            # Default: next to the scripts, then fall back to the output dir.
            here = Path(__file__).resolve().parent / DEFAULT_EXCLUDE_FILENAME
            exclude_path = here if here.exists() else outdir / DEFAULT_EXCLUDE_FILENAME
        file_patterns = load_exclude_patterns(exclude_path)
        if file_patterns:
            print(f"Läste {len(file_patterns)} exkluderingsmönster från {exclude_path}.")
        patterns.extend(file_patterns)
    patterns.extend(args.exclude or [])
    return patterns


def main() -> None:
    p = argparse.ArgumentParser(description="Bygger schemagränssnittet från datafilen.")
    p.add_argument("--input", action="append", default=None, metavar="FIL",
                   help="Datafil (.json/.csv). Kan upprepas. Standard: leta i --outdir.")
    p.add_argument("--outdir", default="medeltidsveckan_output")
    p.add_argument("--slot-minutes", type=int, default=30,
                   help="Tidsupplösning för rutnätets etiketter.")
    p.add_argument("--no-inofficial", action="store_true",
                   help="Ta inte med det inofficiella programmet (imtv.se) även om datafilen finns.")
    p.add_argument("--exclude", action="append", default=[], metavar="TITEL",
                   help="Titel att filtrera bort (kan upprepas). Stöder * och ? som jokertecken.")
    p.add_argument("--exclude-file", default=None,
                   help=f"Fil med titlar att filtrera bort (en per rad). Standard: {DEFAULT_EXCLUDE_FILENAME} bredvid skripten.")
    p.add_argument("--no-exclude-file", action="store_true",
                   help="Ignorera exclude-filen även om den finns.")
    p.add_argument("--no-excel", action="store_true", help="Hoppa över Excel-utdata.")
    p.add_argument("--ics-endpoint", default=None, metavar="URL",
                   help="Bas-URL till prenumerations-Workern, t.ex. "
                        "https://mv.<konto>.workers.dev. Sparas i "
                        f"{CONFIG_FILENAME} och aktiverar Prenumerera-knappen.")
    p.add_argument("--no-ics-endpoint", action="store_true",
                   help="Dölj Prenumerera-knappen även om en endpoint är sparad.")
    p.add_argument("--worker-data", default=None, metavar="FIL",
                   help="Var programdatan för Workern skrivs "
                        "(standard: worker/src/programme.js bredvid skripten).")
    args = p.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    input_paths = _resolve_inputs(outdir, args.input, include_inofficial=not args.no_inofficial)

    events: list[Event] = []
    per_file: list[tuple[Path, int]] = []
    for ip in input_paths:
        loaded = load_events(ip)
        per_file.append((ip, len(loaded)))
        events.extend(loaded)
    events = dedupe(events)
    if not events:
        raise SystemExit(f"Datafilerna innehåller inga event: {', '.join(str(p) for p in input_paths)}")

    patterns = _collect_exclude_patterns(args, outdir)
    if patterns:
        events, removed = filter_events(events, patterns)
        total_removed = sum(removed.values())
        if total_removed:
            print(f"Filtrerade bort {total_removed} event via {len(removed)} titelmönster:")
            for pat, n in removed.most_common():
                print(f"  − {n:3} × {pat}")
        else:
            print("Inga event matchade exkluderingsmönstren.")
        if not events:
            raise SystemExit("Alla event filtrerades bort – kontrollera dina exclude-mönster.")

    config = _load_config(outdir)
    if args.no_ics_endpoint:
        ics_endpoint = ""
    elif args.ics_endpoint is not None:
        ics_endpoint = args.ics_endpoint.rstrip("/")
        config["ics_endpoint"] = ics_endpoint
        _save_config(outdir, config)
        print(f"Sparade prenumerations-endpoint i {outdir / CONFIG_FILENAME}")
    else:
        ics_endpoint = config.get("ics_endpoint", "")

    html_path = outdir / "medeltidsveckan_schema.html"
    write_html(events, html_path, args.slot_minutes, ics_endpoint)
    if len(per_file) == 1:
        print(f"Läste {per_file[0][1]} event från {per_file[0][0]}.")
    else:
        print(f"Läste {len(events)} event (efter sammanslagning) från {len(per_file)} filer:")
        for path, n in per_file:
            print(f"  • {n:4} × {path.name}")
    print(f"Skapade: {html_path}")

    worker_data_path = (
        Path(args.worker_data) if args.worker_data
        else Path(__file__).resolve().parent / "worker" / "src" / "programme.js"
    )
    n_prog = write_worker_data(events, worker_data_path)
    print(f"Skapade: {worker_data_path} ({n_prog} event för prenumeration)")
    if ics_endpoint:
        print(f"Prenumerations-endpoint: {ics_endpoint}")

    if not args.no_excel:
        xlsx_path = outdir / "medeltidsveckan_schema.xlsx"
        write_excel(events, xlsx_path, args.slot_minutes)
        print(f"Skapade: {xlsx_path}")


if __name__ == "__main__":
    main()
